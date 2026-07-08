from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import json
import io
import os
import requests
from datetime import datetime

app = Flask(__name__)
standards_store = {}

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

def parse_excel(file):
    df = pd.read_excel(file, header=None)
    header_row = -1
    std_col = name_col = date_col = -1
    for i in range(min(len(df), 10)):
        row = [str(c).upper() for c in df.iloc[i]]
        for j, cell in enumerate(row):
            if 'STANDART NO' in cell or 'STANDARD NO' in cell:
                header_row = i
                std_col = j
                break
        if header_row >= 0:
            break
    if header_row < 0:
        raise ValueError('Standart NO sütunu bulunamadı')
    headers = [str(c).upper() for c in df.iloc[header_row]]
    for j, h in enumerate(headers):
        if 'STANDART ADI' in h or 'STANDARD NAME' in h:
            name_col = j
        if 'YAYIN' in h or 'REVISION' in h or 'TARİH' in h:
            date_col = j
    records = []
    for i in range(header_row + 1, len(df)):
        row = df.iloc[i]
        no = str(row.iloc[std_col] if std_col >= 0 else '').replace('\xa0', ' ').replace('\n', ' ').strip()
        if not no or len(no) < 3 or no.upper() == 'NAN':
            continue
        name = str(row.iloc[name_col] if name_col >= 0 else '').split('\n')[0].strip()[:100]
        tarih = str(row.iloc[date_col] if date_col >= 0 else '').strip()
        if tarih in ('nan', 'None', 'NaT', ''):
            tarih = ''
        if '00:00:00' in tarih:
            tarih = tarih.replace(' 00:00:00', '').strip()
        records.append({
            'no': no, 'name': name if name != 'nan' else '',
            'tarih': tarih, 'status': 'pending',
            'note': '', 'currentVer': '', 'history': [], 'url': ''
        })
    return records

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya bulunamadı'}), 400
    file = request.files['file']
    try:
        records = parse_excel(file)
        session_id = datetime.now().strftime('%Y%m%d%H%M%S%f')
        standards_store[session_id] = records
        return jsonify({'session_id': session_id, 'count': len(records), 'standards': records})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/check_one', methods=['POST'])
def check_one():
    data = request.json
    session_id = data.get('session_id')
    idx = data.get('idx')
    if session_id not in standards_store:
        return jsonify({'error': 'Oturum bulunamadı'}), 404

    s = standards_store[session_id][idx]
    # Standart kodunu temizle - parantez içi ekleri kaldır
    import re
    clean_no = re.sub(r'\s*\([^)]*\)', '', s['no']).strip()
    up = clean_no.upper()

    if 'ISO' in up:
        src = 'ISO (iso.org)'
    elif 'IEC' in up:
        src = 'IEC (iec.ch)'
    elif up.startswith('TS'):
        src = 'TSE (tse.org.tr)'
    elif 'MDCG' in up:
        src = 'MDCG (ec.europa.eu)'
    elif 'IMDRF' in up or 'GHTF' in up:
        src = 'IMDRF (imdrf.org)'
    elif 'IAF' in up:
        src = 'IAF (iaf.nu)'
    elif up.startswith('EA'):
        src = 'EA (european-accreditation.org)'
    elif 'ILAC' in up:
        src = 'ILAC (ilac.org)'
    else:
        src = 'TSE (tse.org.tr)'

    prompt = f"""Sen bir standartlar uzmanısın. Aşağıdaki standart için web'de arama yaparak güncellik kontrolü yap.

Standart kodu: "{clean_no}"
Standart adı: "{s['name']}"
Bizim listedeki tarih/versiyon: "{s['tarih'] or 'belirtilmemiş'}"
Resmi kaynak: {src}

ÖNEMLİ: Eğer bu standart iptal edilerek yeni bir kodla değiştirildiyse bunu da belirt.

Görevler:
1. Bu standartın TÜM versiyon geçmişini bul
2. En güncel versiyonu belirle  
3. Bizim tarihimizden sonra yeni versiyon var mı?

SADECE JSON formatında yanıt ver, başka hiçbir şey yazma:
{{
  "durum": "güncel" veya "güncelleme_var" veya "iptal_yerine_yeni_var" veya "bulunamadı",
  "guncel_versiyon": "örn: 2024",
  "yerine_gecen": "varsa yeni standart kodu, yoksa null",
  "aciklama": "maksimum 2 cümle Türkçe açıklama",
  "versiyon_gecmisi": [{{"year":"1998","not":"İlk yayın"}},{{"year":"2020","not":"Güncelleme"}}],
  "kaynak_url": "varsa URL, yoksa null"
}}"""

    try:
        res = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'Content-Type': 'application/json',
                'x-api-key': ANTHROPIC_API_KEY,
                'anthropic-version': '2023-06-01'
            },
            json={
                'model': 'claude-haiku-4-5',
                'max_tokens': 600,
                'tools': [{'type': 'web_search_20250305', 'name': 'web_search', 'max_uses': 1}],
                'messages': [{'role': 'user', 'content': prompt}]
            },
            timeout=30
        )

        if not res.ok:
            err = res.json()
            return jsonify({'error': str(err)}), 400

        result = res.json()
        text = ''.join(b['text'] for b in result['content'] if b['type'] == 'text')

        # cite tag'lerini ve sistem etiketlerini temizle
        import re
        text = re.sub(r'<cite[^>]*>', '', text)
        text = re.sub(r'</cite>', '', text)
        text = re.sub(r'<[^>]+>', '', text)

        clean = text.replace('```json', '').replace('```', '').strip()

        # JSON bloğunu metinden çıkar
        json_match = re.search(r'\{[\s\S]*\}', clean)
        if json_match:
            clean = json_match.group(0)
        try:
            parsed = json.loads(clean)
        except:
            parsed = None

        if parsed:
            if parsed.get('durum') in ('güncelleme_var', 'iptal_yerine_yeni_var'):
                status = 'update'
            elif parsed.get('durum') == 'güncel':
                status = 'current'
            else:
                status = 'unknown'

            current_ver = parsed.get('guncel_versiyon', '')
            if parsed.get('yerine_gecen'):
                current_ver = f"Yerine geçen: {parsed['yerine_gecen']}"

            detail = ''
            if parsed.get('yerine_gecen'):
                detail = f"⚠️ Yerine geçen: {parsed['yerine_gecen']} · "
            detail += parsed.get('aciklama', '')

            standards_store[session_id][idx].update({
                'status': status,
                'currentVer': current_ver,
                'note': detail,
                'history': parsed.get('versiyon_gecmisi', []),
                'url': parsed.get('kaynak_url') or ''
            })
            return jsonify({'ok': True, 'standard': standards_store[session_id][idx]})
        else:
            standards_store[session_id][idx]['status'] = 'unknown'
            standards_store[session_id][idx]['note'] = 'Yanıt ayrıştırılamadı'
            return jsonify({'ok': True, 'standard': standards_store[session_id][idx]})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update', methods=['POST'])
def update():
    data = request.json
    session_id = data.get('session_id')
    idx = data.get('idx')
    status = data.get('status')
    note = data.get('note', '')
    if session_id not in standards_store:
        return jsonify({'error': 'Oturum bulunamadı'}), 404
    standards_store[session_id][idx]['status'] = status
    standards_store[session_id][idx]['note'] = note
    return jsonify({'ok': True})

@app.route('/export', methods=['POST'])
def export():
    data = request.json
    session_id = data.get('session_id')
    if session_id not in standards_store:
        return jsonify({'error': 'Oturum bulunamadı'}), 404
    records = standards_store[session_id]
    rows = [['Standart No', 'Standart Adı', 'Listedeki Tarih', 'Durum', 'Güncel Versiyon', 'Açıklama', 'Versiyon Geçmişi']]
    for s in records:
        durum = {'current': 'Güncel', 'update': 'Güncelleme Var', 'unknown': 'Bulunamadı', 'pending': 'Kontrol Edilmedi'}.get(s['status'], '')
        his = ' → '.join(h['year'] + (' (' + h.get('not', '') + ')' if h.get('not') else '') for h in s.get('history', []))
        rows.append([s['no'], s['name'], s['tarih'], durum, s.get('currentVer', ''), s.get('note', ''), his])
    df = pd.DataFrame(rows[1:], columns=rows[0])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sonuçlar')
        ws = writer.sheets['Sonuçlar']
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color='FFE2E2', end_color='FFE2E2', fill_type='solid')
        green_fill = PatternFill(start_color='E2FFE8', end_color='E2FFE8', fill_type='solid')
        for row in ws.iter_rows(min_row=2):
            if row[3].value == 'Güncelleme Var':
                for cell in row: cell.fill = red_fill
            elif row[3].value == 'Güncel':
                for cell in row: cell.fill = green_fill
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    output.seek(0)
    filename = f"standart_kontrol_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
